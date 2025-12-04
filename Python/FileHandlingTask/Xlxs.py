import os
from openpyxl import load_workbook
path=r"E:\Work\Maang-Work\BootCamp\Python\FileHandlingTask"
target="Vaibhav"

for f in os.listdir(path):
    if f.endswith(".xlsx"):
        filename=os.path.join(path,f)
        print("File:",filename)
        cnt=0
        wb=load_workbook(filename)
        for sheet in wb.sheetnames:
            ws=wb[sheet]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell,str):
                        cnt+=cell.count(target)
        print("Count of 'Vaibhav':",cnt)
        print("----------------------------------")
